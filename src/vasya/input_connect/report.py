from dataclasses import dataclass
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
import matplotlib.pyplot as plt
from jinja2 import Environment, FileSystemLoader
import pdfkit
import os
from os import path

from .base import InputConnect
from ..dataset import DataSet

from typing import Any, Dict, List, Union
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell


@dataclass
class StatsData:
    salary: float
    count: int


class InputConnectReport(InputConnect):
    def __init__(self, file_name: str, profession: str) -> None:
        self.file_name = file_name
        self.profession = profession

        self.years_stats: Dict[int, StatsData] = {}
        self.cities_stats: Dict[str, StatsData] = {}
        self.vacancy_stats: Dict[int, StatsData] = {}
        self.total_vacancies = 0

    @classmethod
    def from_input(cls) -> "InputConnectReport":
        file_name = input("Введите название файла: ")
        profession = input("Введите название профессии: ")

        return cls(file_name, profession)

    def prepare_data(self):
        vacancies = DataSet.from_file(self.file_name)
        vacancies.to_list()
        self.count_vacancies(vacancies)
        self.make_stats_as_average()

    def count_vacancies(self, vacancies: DataSet):
        def increment_and_add(data: StatsData, value: float) -> StatsData:
            data.salary += value
            data.count += 1

        self.total_vacancies = len(vacancies)
        for vacancy in vacancies:
            year = vacancy.published_at.year

            self.years_stats.setdefault(year, StatsData(0, 0))
            self.vacancy_stats.setdefault(year, StatsData(0, 0))
            self.cities_stats.setdefault(vacancy.area_name, StatsData(0, 0))

            increment_and_add(self.years_stats[year], vacancy.salary_rub)
            increment_and_add(self.cities_stats[vacancy.area_name], vacancy.salary_rub)

            if self.profession in vacancy.name:
                increment_and_add(self.vacancy_stats[year], vacancy.salary_rub)

    def make_stats_as_average(self):
        for year in self.years_stats:
            self.years_stats[year].salary = int(
                self.years_stats[year].salary // self.years_stats[year].count
            )

        for city in tuple(self.cities_stats):
            percent_count = round(
                self.cities_stats[city].count / self.total_vacancies, 4
            )
            if percent_count < 0.01:
                self.cities_stats.pop(city)
            else:
                self.cities_stats[city].salary = int(
                    self.cities_stats[city].salary // self.cities_stats[city].count
                )
                self.cities_stats[city].count = percent_count

        for year in self.vacancy_stats:
            if self.vacancy_stats[year].count:
                self.vacancy_stats[year].salary = int(
                    self.vacancy_stats[year].salary // self.vacancy_stats[year].count
                )

    def print_answer(self):
        print(
            "Динамика уровня зарплат по годам:",
            {year: self.years_stats[year].salary for year in self.years_stats},
        )
        print(
            "Динамика количества вакансий по годам:",
            {year: self.years_stats[year].count for year in self.years_stats},
        )

        print(
            "Динамика уровня зарплат по годам для выбранной профессии:",
            {year: self.vacancy_stats[year].salary for year in self.vacancy_stats},
        )
        print(
            "Динамика количества вакансий по годам для выбранной профессии:",
            {year: self.vacancy_stats[year].count for year in self.vacancy_stats},
        )

        cities_sorted = sorted(
            self.cities_stats,
            key=lambda x: self.cities_stats[x].salary,
            reverse=True,
        )[:10]
        print(
            "Уровень зарплат по городам (в порядке убывания):",
            {city: self.cities_stats[city].salary for city in cities_sorted},
        )
        cities_sorted = sorted(
            self.cities_stats, key=lambda x: self.cities_stats[x].count, reverse=True
        )[:10]
        print(
            "Доля вакансий по городам (в порядке убывания):",
            {city: self.cities_stats[city].count for city in cities_sorted},
        )

    def get_sorted_cities(self, attr_name: str):
        sorted_names = sorted(
            self.cities_stats,
            key=lambda x: getattr(self.cities_stats[x], attr_name),
            reverse=True,
        )[:10]
        return {name: self.cities_stats[name] for name in sorted_names}

    def get_answer(
        self, template_path: str = "pdf_template.html", *args, **kwargs
    ) -> None:
        self.print_answer()
        Report.generate_excel(self)
        Report.generate_image(self, "graph.png")
        Report.generate_pdf(self, template_path, "report.pdf")


class Report:
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    @classmethod
    def create_header(
        cls, data: Worksheet, letter: str, name: str, width: Union[int, None] = None
    ) -> int:
        cell = f"{letter}1"
        data[cell] = name
        data[cell].font = Font(bold=True)
        data[cell].border = cls.thin_border
        data.column_dimensions[letter].width = result = width or (len(name) + 2)
        return result

    @classmethod
    def set_cell(cls, data: Worksheet, letter: str, row: int, value: Any) -> Cell:
        cell = f"{letter}{row}"
        data[cell] = value
        data[cell].border = cls.thin_border
        return data[cell]

    @classmethod
    def generate_excel(cls, input_connect: InputConnectReport) -> None:
        wb = openpyxl.Workbook()
        wb.remove(wb["Sheet"])

        data = wb.create_sheet("Статистика по годам")
        cls.create_header(data, "A", "Год", 6)
        cls.create_header(data, "B", "Средняя зарплата")
        cls.create_header(data, "C", f"Средняя зарплата - {input_connect.profession}")
        cls.create_header(data, "D", "Количество вакансий")
        cls.create_header(
            data, "E", f"Количество вакансий - {input_connect.profession}"
        )

        for i in input_connect.years_stats:
            cls.set_cell(data, "A", i - 2005, i)
            cls.set_cell(data, "B", i - 2005, input_connect.years_stats[i].salary)
            cls.set_cell(data, "D", i - 2005, input_connect.years_stats[i].count)

        for i in input_connect.vacancy_stats:
            cls.set_cell(data, "C", i - 2005, input_connect.vacancy_stats[i].salary)
            cls.set_cell(data, "E", i - 2005, input_connect.vacancy_stats[i].count)

        data = wb.create_sheet("Статистика по городам")
        len_a = cls.create_header(data, "A", "Город")
        cls.create_header(data, "B", "Уровень зарплат")
        len_d = cls.create_header(data, "D", "Город")
        cls.create_header(data, "E", "Доля вакансий")

        sorted_cities = input_connect.get_sorted_cities("salary")
        for n, i in enumerate(sorted_cities, 2):
            cls.set_cell(data, "A", n, i)
            data.column_dimensions["A"].width = len_a = max(len_a, len(i) + 2)
            cls.set_cell(data, "B", n, sorted_cities[i].salary)

        sorted_cities = input_connect.get_sorted_cities("count")
        for n, i in enumerate(sorted_cities, 2):
            cls.set_cell(data, "D", n, i)
            data.column_dimensions["D"].width = len_d = max(len_d, len(i) + 2)
            cell = cls.set_cell(
                data, "E", n, f"{round(sorted_cities[i].count * 100, 2)}%"
            )
            cell.number_format = "0.00%"

        wb.save("report.xlsx")

    @classmethod
    def add_simple_graph(
        cls,
        axes: plt.Axes,
        x_val: List[int],
        y_val1: List[float],
        y_val2: List[float],
        name_values1: str,
        name_values2: str,
        title: str,
    ) -> None:
        axes.set_title(title, fontsize=16)
        axes.grid(axis="y")
        axes.bar([v + 0.2 for v in x_val], y_val2, label=name_values2, width=0.4)
        axes.bar([v - 0.2 for v in x_val], y_val1, label=name_values1, width=0.4)
        axes.legend()
        axes.tick_params(axis="x", labelrotation=90)

    @classmethod
    def add_horizontal_graph(
        cls, axes: plt.Axes, x_val: List[str], y_val: List[float], title: str
    ) -> None:
        axes.set_title(title, fontsize=16)
        axes.grid(axis="x")
        axes.barh(x_val, y_val)
        axes.invert_yaxis()

    @classmethod
    def add_circle_diagramm(
        cls, axes: plt.Axes, names: List[str], values: List[int], title: str
    ) -> None:
        axes.set_title(title, fontsize=16)
        names.append("Другие")
        values.append(1 - sum(values))
        axes.pie(values, labels=names)

    @classmethod
    def generate_image(cls, input_connect: InputConnectReport, filename: str) -> None:
        fig, axis = plt.subplots(2, 2)
        plt.rcParams["font.size"] = 8
        cls.add_simple_graph(
            axis[0, 0],
            input_connect.years_stats,
            [
                input_connect.years_stats[key].salary
                for key in input_connect.years_stats
            ],
            [
                input_connect.vacancy_stats[key].salary
                for key in input_connect.vacancy_stats
            ],
            "Средняя з/п",
            f"з/п {input_connect.profession}",
            "Уровень зарплат по годам",
        )
        cls.add_simple_graph(
            axis[0, 1],
            input_connect.years_stats,
            [input_connect.years_stats[key].count for key in input_connect.years_stats],
            [
                input_connect.vacancy_stats[key].count
                for key in input_connect.vacancy_stats
            ],
            "Количество вакансий",
            f"Количество вакансий {input_connect.profession}",
            "Количество вакансий по годам",
        )
        sorted_cities_by_salary = input_connect.get_sorted_cities("salary")
        cls.add_horizontal_graph(
            axis[1, 0],
            [key for key in sorted_cities_by_salary],
            [sorted_cities_by_salary[key].salary for key in sorted_cities_by_salary],
            "Уровень зарплат по городам",
        )
        sorted_cities_by_count = input_connect.get_sorted_cities("count")
        cls.add_circle_diagramm(
            axis[1, 1],
            [key for key in sorted_cities_by_count],
            [sorted_cities_by_count[key].count for key in sorted_cities_by_count],
            "Доля вакансий по городам",
        )
        plt.axis("equal")
        fig.set_size_inches(16, 9)
        fig.tight_layout(h_pad=1)
        plt.savefig(filename)

    @classmethod
    def generate_pdf(
        cls, input_connect: InputConnectReport, template_name: str, filename: str
    ) -> None:
        env = Environment(loader=FileSystemLoader("."))
        template = env.get_template(template_name)
        render_rules = cls.get_render_rules(input_connect)
        pdf_template = template.render(render_rules)
        config = pdfkit.configuration(
            wkhtmltopdf=(
                "C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe"
                if os.name == "nt"
                else "/usr/bin/wkhtmltopdf"
            )
        )
        pdfkit.from_string(
            pdf_template,
            filename,
            configuration=config,
            options={"enable-local-file-access": True},
        )

    @classmethod
    def get_render_rules(cls, input_connect: InputConnectReport) -> Dict[str, Any]:
        rules = {
            "profession": input_connect.profession,
            "image_path": path.abspath("graph.png"),
        }

        for i in input_connect.years_stats:
            rules[f"all_avg_salary{i}"] = input_connect.years_stats[i].salary
            rules[f"all_count{i}"] = input_connect.years_stats[i].count
        for i in input_connect.vacancy_stats:
            rules[f"profession_avg_salary{i}"] = input_connect.vacancy_stats[i].salary
            rules[f"profession_count{i}"] = input_connect.vacancy_stats[i].count

        cities = input_connect.get_sorted_cities("salary")
        for n, i in enumerate(cities, start=1):
            rules[f"table1_city{n}"] = i
            rules[f"city_salary{n}"] = cities[i].salary

        cities = input_connect.get_sorted_cities("count")
        for n, i in enumerate(cities, start=1):
            rules[f"table2_city{n}"] = i
            rules[f"city_count{n}"] = f"{round(cities[i].count * 100, 2)}%"

        return rules