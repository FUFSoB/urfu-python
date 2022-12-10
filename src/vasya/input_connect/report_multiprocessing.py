from dataclasses import dataclass
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
import matplotlib.pyplot as plt
from jinja2 import Environment, FileSystemLoader
import pdfkit
import os
from os import path
from pathlib import Path
from multiprocessing import Process, Manager as MPManager

from .base import InputConnect
from ..dataset import DataSet

from typing import Any, Dict, List, Optional, Tuple
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell


@dataclass
class StatsData:
    """
    Класс для хранения данных для отчёта.
    """

    salary: float
    count: int

    def copy(self) -> "StatsData":
        """
        Метод для копирования объекта.

        Returns
        -------
        StatsData
            Копия объекта
        """

        return StatsData(self.salary, self.count)


class InputConnectReportMultiprocessing(InputConnect):
    """
    Класс-коннектор для создания отчёта.

    Attributes
    ----------
    dir_name: Path
        Путь до директории с csv файлами
    profession: str
        Профессия, по которой будет производиться анализ
    years_stats: Dict[int, StatsData]
        Статистика по годам
    cities_stats: Dict[str, StatsData]
        Статистика по городам
    vacancy_stats: Dict[int, StatsData]
        Статистика по годам по профессии
    total_vacancies: int
        Общее количество вакансий
    """

    ReturnType = Tuple[StatsData, Dict[str, StatsData], StatsData]

    def __init__(self, dir_name: str, profession: str) -> None:
        """
        Инициализация класса

        Parameters
        ----------
        dir_name: str
            Путь до директории с csv файлами
        profession: str
            Профессия, по которой будет производиться анализ
        """
        self.dir_name = Path(dir_name)
        self.profession = profession

        self.years_stats: Dict[int, StatsData] = {}
        self.cities_stats: Dict[str, StatsData] = {}
        self._proc_cities_stats: List[Dict[str, StatsData]] = []
        self.vacancy_stats: Dict[int, StatsData] = {}
        self.total_vacancies = 0

    @classmethod
    def from_input(cls) -> "InputConnectReportMultiprocessing":
        """
        Метод для создания объекта класса по вводу пользователя.

        Returns
        -------
        InputConnectReportMultiprocessing
            Объект класса
        """
        dir_name = input("Введите путь до директории с файлами csv: ")
        profession = input("Введите название профессии: ")

        return cls(dir_name, profession)

    def prepare_data(self) -> None:
        """
        Метод для подготовки данных для отчёта.

        Raises
        ------
        VasyaException
            Файл не найден или в нём нет данных
        """

        manager = MPManager()
        return_dict: Dict[int, self.ReturnType] = manager.dict()

        processes: List[Process] = []

        for file_name in self.dir_name.glob("*.csv"):
            try:
                year = int(file_name.stem.rsplit("_", 1)[1])
            except ValueError:
                print(
                    f"Пропускаем файл: {file_name} - "
                    "должен быть в формате name_year.csv (например: data_2020.csv)"
                )
                continue

            process = Process(
                target=self.process_file, args=(file_name, year, return_dict)
            )
            processes.append(process)

        for process in processes:
            process.start()

        for process in processes:
            process.join()

        for year, (years_stats, cities_stats, vacancy_stats) in return_dict.items():
            self.years_stats[year] = years_stats
            self._proc_cities_stats.append(cities_stats)
            self.vacancy_stats[year] = vacancy_stats

        self.make_stats_as_average()

    def process_file(
        self, file_name: Path, year: int, return_dict: Dict[int, ReturnType]
    ) -> None:
        """
        Метод для обработки файла.

        Parameters
        ----------
        file_name: Path
            Путь до файла
        year: int
            Год, за который собраны данные
        return_dict: Dict[int, ReturnType]
            Словарь для возврата данных

        Raises
        ------
        VasyaException
            Файл не найден, в нём нет данных или неверный формат имени файла
        """

        years_stats = StatsData(0, 0)
        cities_stats: Dict[str, StatsData] = {}
        vacancy_stats = StatsData(0, 0)

        vacancies = DataSet.from_file(file_name)
        vacancies.to_list()

        self._count_vacancies(vacancies, years_stats, cities_stats, vacancy_stats)
        return_dict[year] = (years_stats, cities_stats, vacancy_stats)

    def _count_vacancies(
        self,
        vacancies: DataSet,
        years_stats: StatsData,
        cities_stats: Dict[str, StatsData],
        vacancy_stats: StatsData,
    ) -> None:
        """
        Метод для подсчёта статистики по вакансиям.

        Parameters
        ----------
        vacancies: DataSet
            Датасет с вакансиями
        years_stats: StatsData
            Статистика по году
        cities_stats: Dict[str, StatsData]
            Статистика по городам
        vacancy_stats: StatsData
            Статистика по вакансиям
        """

        def increment_and_add(data: StatsData, value: float) -> StatsData:
            data.salary += value
            data.count += 1

        for vacancy in vacancies:
            cities_stats.setdefault(vacancy.area_name, StatsData(0, 0))

            increment_and_add(years_stats, vacancy.salary_rub)
            increment_and_add(cities_stats[vacancy.area_name], vacancy.salary_rub)

            if self.profession in vacancy.name:
                increment_and_add(vacancy_stats, vacancy.salary_rub)

    def make_stats_as_average(self) -> None:
        """
        Метод для конвертации существующей статистики в среднюю зарплату.
        """

        for year in self.years_stats:
            self.years_stats[year].salary = int(
                self.years_stats[year].salary // self.years_stats[year].count
            )
            self.total_vacancies += self.years_stats[year].count

        for cities_stats in self._proc_cities_stats:
            for key, value in cities_stats.items():
                if self.cities_stats.get(key):
                    self.cities_stats[key].salary += value.salary
                    self.cities_stats[key].count += value.count
                else:
                    self.cities_stats[key] = value.copy()

        for city in tuple(self.cities_stats):
            percent_count = round(
                self.cities_stats[city].count / self.total_vacancies, 4
            )
            if percent_count < 0.01:
                self.cities_stats.pop(city)
            else:
                self.cities_stats[city].salary = int(
                    self.cities_stats[city].salary // self.cities_stats[city].count
                )
                self.cities_stats[city].count = percent_count

        for year in self.vacancy_stats:
            if self.vacancy_stats[year].count:
                self.vacancy_stats[year].salary = int(
                    self.vacancy_stats[year].salary // self.vacancy_stats[year].count
                )

    def print_answer(self) -> None:
        """
        Метод для вывода ответа в консоль.
        """

        print(
            "Динамика уровня зарплат по годам:",
            {year: self.years_stats[year].salary for year in self.years_stats},
        )
        print(
            "Динамика количества вакансий по годам:",
            {year: self.years_stats[year].count for year in self.years_stats},
        )

        print(
            "Динамика уровня зарплат по годам для выбранной профессии:",
            {year: self.vacancy_stats[year].salary for year in self.vacancy_stats},
        )
        print(
            "Динамика количества вакансий по годам для выбранной профессии:",
            {year: self.vacancy_stats[year].count for year in self.vacancy_stats},
        )

        cities_sorted = sorted(
            self.cities_stats,
            key=lambda x: self.cities_stats[x].salary,
            reverse=True,
        )[:10]
        print(
            "Уровень зарплат по городам (в порядке убывания):",
            {city: self.cities_stats[city].salary for city in cities_sorted},
        )
        cities_sorted = sorted(
            self.cities_stats, key=lambda x: self.cities_stats[x].count, reverse=True
        )[:10]
        print(
            "Доля вакансий по городам (в порядке убывания):",
            {city: self.cities_stats[city].count for city in cities_sorted},
        )

    def get_sorted_cities(self, attr_name: str) -> Dict[str, StatsData]:
        """
        Метод для получения отсортированных городов по атрибуту.

        Parameters
        ----------
        attr_name: str
            Название атрибута

        Returns
        -------
        Dict[str, StatsData]
            Словарь из десяти городов, отсортированных по атрибуту
        """

        sorted_names = sorted(
            self.cities_stats,
            key=lambda x: getattr(self.cities_stats[x], attr_name),
            reverse=True,
        )[:10]
        return {name: self.cities_stats[name] for name in sorted_names}

    def get_answer(
        self, template_path: str = "pdf_template.html", *args, **kwargs
    ) -> None:
        """
        Метод для получения результата обработки входных данных.

        Выводит статистику в консоль.

        Создаёт файлы report.xlsx, graph.png и report.pdf
        в директории, из которой был запущен скрипт.

        Parameters
        ----------
        template_path: str
            Путь к шаблону для генерации pdf-файла
        """

        self.print_answer()
        Report.generate_excel(self, "report.xlsx")
        Report.generate_image(self, "graph.png")
        Report.generate_pdf(self, template_path, "report.pdf")


class Report:
    """
    Класс. хранящий в себе статические методы для генерации отчётов.
    """

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    @classmethod
    def create_header(
        cls, data: Worksheet, letter: str, name: str, width: Optional[int] = None
    ) -> int:
        """
        Метод для создания заголовка столбца.

        Parameters
        ----------
        data: Worksheet
            Лист, в котором будет создан заголовок
        letter: str
            Буква столбца
        name: str
            Название заголовка
        width: Optional[int]
            Ширина столбца

        Returns
        -------
        int
            Ширина столбца
        """

        cell = f"{letter}1"
        data[cell] = name
        data[cell].font = Font(bold=True)
        data[cell].border = cls.thin_border
        data.column_dimensions[letter].width = result = width or (len(name) + 2)
        return result

    @classmethod
    def set_cell(cls, data: Worksheet, letter: str, row: int, value: Any) -> Cell:
        """
        Метод для создания ячейки.

        Parameters
        ----------
        data: Worksheet
            Лист, в котором будет создана ячейка
        letter: str
            Буква столбца
        row: int
            Номер строки
        value: Any
            Значение ячейки

        Returns
        -------
        Cell
            Созданная ячейка
        """

        cell = f"{letter}{row}"
        data[cell] = value
        data[cell].border = cls.thin_border
        return data[cell]

    @classmethod
    def generate_excel(
        cls, input_connect: InputConnectReportMultiprocessing, filename: str
    ) -> None:
        """
        Метод для генерации excel-файла.
        Создаёт файл filename в директории, из которой был запущен скрипт.

        Parameters
        ----------
        input_connect: InputConnectReportMultiprocessing
            Объект, хранящий в себе статистику
        filename: str
            Название файла
        """

        wb = openpyxl.Workbook()
        wb.remove(wb["Sheet"])

        data = wb.create_sheet("Статистика по годам")
        cls.create_header(data, "A", "Год", 6)
        cls.create_header(data, "B", "Средняя зарплата")
        cls.create_header(data, "C", f"Средняя зарплата - {input_connect.profession}")
        cls.create_header(data, "D", "Количество вакансий")
        cls.create_header(
            data, "E", f"Количество вакансий - {input_connect.profession}"
        )

        for i in input_connect.years_stats:
            cls.set_cell(data, "A", i - 2005, i)
            cls.set_cell(data, "B", i - 2005, input_connect.years_stats[i].salary)
            cls.set_cell(data, "D", i - 2005, input_connect.years_stats[i].count)

        for i in input_connect.vacancy_stats:
            cls.set_cell(data, "C", i - 2005, input_connect.vacancy_stats[i].salary)
            cls.set_cell(data, "E", i - 2005, input_connect.vacancy_stats[i].count)

        data = wb.create_sheet("Статистика по городам")
        len_a = cls.create_header(data, "A", "Город")
        cls.create_header(data, "B", "Уровень зарплат")
        len_d = cls.create_header(data, "D", "Город")
        cls.create_header(data, "E", "Доля вакансий")

        sorted_cities = input_connect.get_sorted_cities("salary")
        for n, i in enumerate(sorted_cities, 2):
            cls.set_cell(data, "A", n, i)
            data.column_dimensions["A"].width = len_a = max(len_a, len(i) + 2)
            cls.set_cell(data, "B", n, sorted_cities[i].salary)

        sorted_cities = input_connect.get_sorted_cities("count")
        for n, i in enumerate(sorted_cities, 2):
            cls.set_cell(data, "D", n, i)
            data.column_dimensions["D"].width = len_d = max(len_d, len(i) + 2)
            cell = cls.set_cell(
                data, "E", n, f"{round(sorted_cities[i].count * 100, 2)}%"
            )
            cell.number_format = "0.00%"

        wb.save(filename)

    @classmethod
    def add_simple_graph(
        cls,
        axes: plt.Axes,
        x_val: List[int],
        y_val1: List[float],
        y_val2: List[float],
        name_values1: str,
        name_values2: str,
        title: str,
    ) -> None:
        """
        Метод для добавления простого графика.

        Parameters
        ----------
        axes: plt.Axes
            Объект, хранящий в себе данные для построения графика
        x_val: List[int]
            Список значений по оси X
        y_val1: List[float]
            Список значений по оси Y для первого графика
        y_val2: List[float]
            Список значений по оси Y для второго графика
        name_values1: str
            Название первого графика
        name_values2: str
            Название второго графика
        title: str
            Название графика
        """

        axes.set_title(title, fontsize=16)
        axes.grid(axis="y")
        axes.bar([v + 0.2 for v in x_val], y_val2, label=name_values2, width=0.4)
        axes.bar([v - 0.2 for v in x_val], y_val1, label=name_values1, width=0.4)
        axes.legend()
        axes.tick_params(axis="x", labelrotation=90)

    @classmethod
    def add_horizontal_graph(
        cls, axes: plt.Axes, x_val: List[str], y_val: List[float], title: str
    ) -> None:
        """
        Метод для добавления горизонтального графика.

        Parameters
        ----------
        axes: plt.Axes
            Объект, хранящий в себе данные для построения графика
        x_val: List[str]
            Список значений по оси X
        y_val: List[float]
            Список значений по оси Y
        title: str
            Название графика
        """

        axes.set_title(title, fontsize=16)
        axes.grid(axis="x")
        axes.barh(x_val, y_val)
        axes.invert_yaxis()

    @classmethod
    def add_circle_diagramm(
        cls, axes: plt.Axes, names: List[str], values: List[int], title: str
    ) -> None:
        """
        Метод для добавления круговой диаграммы.

        Parameters
        ----------
        axes: plt.Axes
            Объект, хранящий в себе данные для построения графика
        names: List[str]
            Список названий секторов
        values: List[int]
            Список значений секторов
        title: str
            Название графика
        """

        axes.set_title(title, fontsize=16)
        names.append("Другие")
        values.append(1 - sum(values))
        axes.pie(values, labels=names)

    @classmethod
    def generate_image(
        cls, input_connect: InputConnectReportMultiprocessing, filename: str
    ) -> None:
        """
        Метод для генерации изображения.
        Создаёт файл filename в директории, из которой был запущен скрипт.

        Parameters
        ----------
        input_connect: InputConnectReportMultiprocessing
            Объект, хранящий в себе данные для построения графика
        filename: str
            Название файла
        """

        fig, axis = plt.subplots(2, 2)
        plt.rcParams["font.size"] = 8
        cls.add_simple_graph(
            axis[0, 0],
            input_connect.years_stats,
            [
                input_connect.years_stats[key].salary
                for key in input_connect.years_stats
            ],
            [
                input_connect.vacancy_stats[key].salary
                for key in input_connect.vacancy_stats
            ],
            "Средняя з/п",
            f"з/п {input_connect.profession}",
            "Уровень зарплат по годам",
        )
        cls.add_simple_graph(
            axis[0, 1],
            input_connect.years_stats,
            [input_connect.years_stats[key].count for key in input_connect.years_stats],
            [
                input_connect.vacancy_stats[key].count
                for key in input_connect.vacancy_stats
            ],
            "Количество вакансий",
            f"Количество вакансий {input_connect.profession}",
            "Количество вакансий по годам",
        )
        sorted_cities_by_salary = input_connect.get_sorted_cities("salary")
        cls.add_horizontal_graph(
            axis[1, 0],
            [key for key in sorted_cities_by_salary],
            [sorted_cities_by_salary[key].salary for key in sorted_cities_by_salary],
            "Уровень зарплат по городам",
        )
        sorted_cities_by_count = input_connect.get_sorted_cities("count")
        cls.add_circle_diagramm(
            axis[1, 1],
            [key for key in sorted_cities_by_count],
            [sorted_cities_by_count[key].count for key in sorted_cities_by_count],
            "Доля вакансий по городам",
        )
        plt.axis("equal")
        fig.set_size_inches(16, 9)
        fig.tight_layout(h_pad=1)
        plt.savefig(filename)

    @classmethod
    def generate_pdf(
        cls,
        input_connect: InputConnectReportMultiprocessing,
        template_name: str,
        filename: str,
    ) -> None:
        """
        Метод для генерации pdf-файла.
        Создаёт файл filename в директории, из которой был запущен скрипт.

        Parameters
        ----------
        input_connect: InputConnectReportMultiprocessing
            Объект, хранящий в себе данные для построения графика
        template_name: str
            Путь до файла шаблона
        filename: str
            Название файла
        """

        env = Environment(loader=FileSystemLoader("."))
        template = env.get_template(template_name)
        render_rules = cls.get_render_rules(input_connect)
        pdf_template = template.render(render_rules)
        config = pdfkit.configuration(
            wkhtmltopdf=(
                "C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe"
                if os.name == "nt"
                else "/usr/bin/wkhtmltopdf"
            )
        )
        pdfkit.from_string(
            pdf_template,
            filename,
            configuration=config,
            options={"enable-local-file-access": True},
        )

    @classmethod
    def get_render_rules(
        cls, input_connect: InputConnectReportMultiprocessing
    ) -> Dict[str, Any]:
        """
        Метод для получения правил отрисовки шаблона.

        Parameters
        ----------
        input_connect: InputConnectReportMultiprocessing
            Объект, хранящий в себе данные для построения графика

        Returns
        -------
        Dict[str, Any]
            Словарь с правилами отрисовки
        """

        rules = {
            "profession": input_connect.profession,
            "image_path": path.abspath("graph.png"),
        }

        for i in input_connect.years_stats:
            rules[f"all_avg_salary{i}"] = input_connect.years_stats[i].salary
            rules[f"all_count{i}"] = input_connect.years_stats[i].count
        for i in input_connect.vacancy_stats:
            rules[f"profession_avg_salary{i}"] = input_connect.vacancy_stats[i].salary
            rules[f"profession_count{i}"] = input_connect.vacancy_stats[i].count

        cities = input_connect.get_sorted_cities("salary")
        for n, i in enumerate(cities, start=1):
            rules[f"table1_city{n}"] = i
            rules[f"city_salary{n}"] = cities[i].salary

        cities = input_connect.get_sorted_cities("count")
        for n, i in enumerate(cities, start=1):
            rules[f"table2_city{n}"] = i
            rules[f"city_count{n}"] = f"{round(cities[i].count * 100, 2)}%"

        return rules
